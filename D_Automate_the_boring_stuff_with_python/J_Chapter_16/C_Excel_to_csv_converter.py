"""Este Programa conviete a csv todos los archivos dee excel que se encuentren en un directorio"""

import openpyxl,csv
from os import listdir,path
from openpyxl.utils.cell import get_column_letter

paths= 'INSERT_PATH_TO_FILE_HERE'

for file in listdir(paths):
    #Avoid non .xlsx
    if not file.endswith('.xlsx'):
        continue
    #open the first .xlsx file it finds
    wb=openpyxl.load_workbook(f'{paths}/{file}')
    #looks for every sheet in the workbook
    for sheet_name in wb.sheetnames:
        sheet=wb[sheet_name]
        #name of the csv file
        csv_file_name=f'{file.strip(".xlsx")}_{sheet_name}.csv'
        #creating a direction to save the csv file
        csv_file_obj=open(path.join("CC_converted_files",csv_file_name),'w',newline='')
        csv_writer=csv.writer(csv_file_obj)
        #go row for row
        for row_num in range(1,sheet.max_row+1):
            #here it keeps all the contents in a row
            row_data=[]
            for column_num in range(1,sheet.max_column+1):
               #iwth the row number and the column number que the vauee of a cell and add it to the list
               row_data.append(sheet[f"{get_column_letter(column_num)}{row_num}"].value)
            #writes the list in the csv file
            csv_writer.writerow(row_data)
        #once it has ende with that sheet continues with the other
        csv_file_obj.close()
    # once it has finished with this workbook it closes it an start with the other
    wb.close()




