"""Este programa manda un recordatorio de pago a los clientes que se encuentra en una lista en excel"""

import openpyxl, smtplib

#opening workbook

wb=openpyxl.load_workbook('BB_Dues_records.xlsx')
sheet=wb["Sheet1"]

ultima_col=sheet.max_column
ultimo_mes=sheet.cell(row=1,column=ultima_col).value

pago_atrasado={}
for row in range(2,sheet.max_row+1):
    pago=sheet.cell(row=row,column=ultima_col).value
    if pago!='paid':
        nombre=sheet.cell(row=row,column=1).value
        correo=sheet.cell(row=row,column=2).value
        pago_atrasado[nombre]=[correo]

#Logeando la cuanta de e_mail.

smtp_obj= smtplib.SMTP('dominio',0)#esto es el nombre del dominio del servidor y el numero del puerto
smtp_obj.ehlo()
smtp_obj.starttls()
smtp_obj.login('mi_correo','mi_contrasenia')
 # abriendo el diccionario donde se encuentran los clientes que no han pagado.
for nombre,correo in pago_atrasado.items():
    cuerpo_email=f'Subject:{ultimo_mes} pago atrasao.\nEstimado {nombre},\nLe escribimos para recordarle que no ' \
                 f'ha pagado el mes de {ultimo_mes} favor realizar el pago lo mas antes posible.'
    print(f'Enviando email a: {correo}')
    email_status=smtp_obj.sendmail('mi_correo',correo,cuerpo_email)

    if email_status!={}:
        print(f'Ha ocurrido un problema para enviar el correo a la direccion {correo}: {email_status}')

smtp_obj.quit()
