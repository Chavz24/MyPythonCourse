"""Este programa cambia la posicion de los valores de la celdas en un documento
    de tal manera que el valor de la celda A2 pasee a se r el valor de la celda
        B1 y viceversa"""
import openpyxl, logging
from openpyxl.utils.cell import get_column_letter,coordinate_to_tuple

#open('mylog.txt','w').write('\n')
logging.basicConfig(filename='mylog.txt',
                    level=logging.DEBUG,
                    format='%(asctime)s - %(levelname)s - %(message)s')
logging.disable(logging.CRITICAL)

#Esto lo use para crear la tabla con los valores
# wb = openpyxl.Workbook()
#
# sheet = wb['Sheet']
# list = ['','Nombre', 'Rosa', 'Fulano', 'Maria', 'Tramontina', 'Gula']
# list2 = ['','Direccion', '234st', 'rcvst', 'boav', 't56st', 'fan451']
# list3 = ['','Telefono', 3567, 55677, 1234, 123234, 45678]
# for i in range(1,len(list)):
#     sheet[f'A{i}'] = list[i]
#     sheet[f'B{i}'] = list2[i]
#     sheet[f'C{i}'] = list3[i]
#
# wb.save('FF_Cell_inverter.xlsx')

wb=openpyxl.load_workbook("FF_Cell_inverter.xlsx")
sheet=wb["Sheet"]

wb2=openpyxl.Workbook()
sheet1=wb2['Sheet']
sheet1.title='Chavez'


min_row=sheet.min_row
max_row=sheet.max_row
min_col=sheet.min_column
max_col=sheet.max_column

logging.info(f'{min_col}  {min_row}  {max_col}  {max_row}')

for rows in sheet[f'{get_column_letter(min_col)}{min_row}':
                  f'{get_column_letter(max_col)}{max_row}']:
    logging.info(f'{rows}')
    for cells in rows:
        coords=(coordinate_to_tuple(cells.coordinate))
        sheet1[f'{get_column_letter(coords[0])}{coords[1]}']=sheet[cells.coordinate].value

wb2.save('FF_Cell_inverterted.xlsx')



