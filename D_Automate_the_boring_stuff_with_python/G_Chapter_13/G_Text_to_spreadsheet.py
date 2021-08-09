"""Este programa recorre un directorio en busca de archivos de texto e imprime su
    nombre een la primera fila de la columna y el resto de su contenido en las filas
    posteriores de la misma columna,por cada archivo que encuentre imprime su sontenido
    en columnas diferentes"""

import openpyxl,logging,os
from openpyxl.utils.cell import get_column_letter
from pathlib import Path

open("mylog.txt",'w').write('\n')
logging.basicConfig(filename='mylog.txt',
                    level=logging.DEBUG,
                    format=f'%(asctime)s - %(levelname)s - %(message)s')
#logging.disable(logging.CRITICAL)

wb=openpyxl.Workbook()

sheet=wb['Sheet']

working_folder=Path.cwd()
width=0
count=1
for file in working_folder.glob('*.txt'):
    file_name=os.path.basename(file)
    my_text=open(f'{file_name}','r')
    lines=my_text.readlines()
    sheet[f'{get_column_letter(count)}1']=file_name
    for line in lines:
        width=max(width,len(line))
        if line=='\n':
            continue
        sheet[f'{get_column_letter(count)}{lines.index(line)+2}']=line
        sheet.column_dimensions[f'{get_column_letter(count)}'].width=width
    count+=1

wb.save('GG_Text_to_spradsheet.xlsx')
