

import openpyxl
from openpyxl.utils.cell import get_column_letter

wb=openpyxl.load_workbook('HH_Spreadsheet_to_text.xlsx')

sheet=wb['Sheet']

column_num=1
count=0
while column_num<=sheet.max_column:
    for rows in sheet[f'{get_column_letter(column_num)}{sheet.min_row}':
                      f'{get_column_letter(column_num)}{sheet.max_row}']:
        for text in rows:
            file_name=f'{sheet[f"{get_column_letter(column_num)}1"].value}'
            if count!=0:
                text_file = open(f'HH_{file_name}', 'a')
                text_file.write(f'{text.value}')
                text_file.close()
            else:
                count+=1


    count=0
    column_num+=1
wb.close()
