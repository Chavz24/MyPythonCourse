#Extrayendo informacion del censo 2010
import openpyxl, pprint

print('Abriendo libro de trabajo...')

wb = openpyxl.load_workbook('censuspopdata.xlsx')
countydata={}
sheet = wb[wb.sheetnames[0]]

for row in range(2, sheet.max_row + 1):
    state   = sheet['B' + str(row)].value
    county  = sheet['C' + str(row)].value
    pop     = sheet['D' + str(row)].value

    countydata.setdefault(state,{})
    countydata[state].setdefault(county,{'tracts':0,'pop':0})

    countydata[state][county]['tracts']+=1
    countydata[state][county]['pop']+=int(pop)

print('Los resultados del censo son...')

resultados=open('censo2010.py','w')
resultados.write('datos_censo='+ pprint.pformat(countydata))
resultados.close()
print('Terminado')
