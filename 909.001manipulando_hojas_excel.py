#Sustituyendo precios en un documento de excel

import openpyxl

wb=openpyxl.load_workbook('produceSales.xlsx')
sheet=wb['Sheet']

price_update={'Garlic': 3.07,
              'Celery': 1.19,
              'Lemon': 1.27}

for row_num in range(2,sheet.max_row):
    nombre_producto=sheet.cell(row=row_num,column=1).value
    if nombre_producto in price_update:
        sheet.cell(row=row_num,column=2).value= price_update[nombre_producto]

wb.save('updated.xlsx')