"""Este programa imprime una tabla de multiplicar en excel"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import  Font

wb=openpyxl.Workbook()


sheet=wb["Sheet"]


for i in range(1,14):
    font=Font(bold=True)
    sheet[f'A{i+1}'].font=font
    sheet[f'{get_column_letter(i + 1)}1'].font=font
    sheet.freeze_panes='B2'


    sheet[f'A{i+1}']=i
    sheet[f'{get_column_letter(i+1)}1']=i
    column_letter=get_column_letter(i+1)
    row_num=i+1


    for j in range(2,15):
        sheet[f'{column_letter}{j}'] = f'={column_letter}1*A{j}'


wb.save("DD_Multiplication_table.xlsx")